"""
Find the observations column in Section 3
Day 9 should have "DIA DA FAMÍLIA NA ESCOLA"
"""

import openpyxl

excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['EMEI']

day9_row = 85  # Day 9

print("="*100)
print("FINDING OBSERVATIONS COLUMN")
print("="*100)

print(f"\nScanning Day 9 (Row {day9_row}) for observation text...")
for col_idx in range(1, 30):
    cell_val = ws.cell(day9_row, col_idx).value
    if cell_val and isinstance(cell_val, str) and len(cell_val) > 3:
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): '{cell_val}'")

# Check day 29 too (user mentioned "IQEIP")
day29_row = 105  # Day 29 (77 + 28)
print(f"\nScanning Day 29 (Row {day29_row}) for observation text...")
for col_idx in range(1, 30):
    cell_val = ws.cell(day29_row, col_idx).value
    if cell_val and isinstance(cell_val, str) and len(cell_val) > 1:
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): '{cell_val}'")

print("\n" + "="*100)
print("COMPLETE SECTION 3 COLUMN MAPPING")
print("="*100)
print("""
Based on the data:

Excel Column -> Field Name -> Example Values
---------------------------------------------------------------------------
Col  3 (C)  -> Dia (Day number) -> 1, 2, 3, ..., 31, Total
Col  4 (D)  -> Grupo_A_Frequencia -> (empty in this file, Total=0)
Col  6 (F)  -> Grupo_A_Lanche_4H -> (empty, Total=0)
Col  8 (H)  -> Grupo_A_Lanche_6H -> (empty, Total=0)
Col 11 (K)  -> Grupo_A_Refeicao_Enteral -> (empty, Total=0)
Col 13 (M)  -> Grupo_B_Frequencia -> Day1=16, Day9=4, Total=332
Col 15 (O)  -> Grupo_B_Lanche_4H -> (empty, Total=0)
Col 17 (Q)  -> Grupo_B_Lanche_6H -> Day1=16, Day9=4, Total=332
Col 19 (S)  -> Lanche_Emergencial -> (empty, Total=0)
Col 21 (U)  -> Kit_Lanche -> (empty, Total=0)
Col 23+ -> Observacoes -> Day9="DIA DA FAMÍLIA...", Day29="IQEIP"
""")