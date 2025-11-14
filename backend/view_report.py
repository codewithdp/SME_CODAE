"""
Quick script to view the reconciliation report summary
"""
import openpyxl
import sys

if len(sys.argv) < 2:
    print("Usage: python view_report.py <report_file.xlsx>")
    sys.exit(1)

report_path = sys.argv[1]

wb = openpyxl.load_workbook(report_path, data_only=True)
ws = wb['Summary']

print("=" * 100)
print("RECONCILIATION REPORT SUMMARY")
print("=" * 100)
print()

# Read key information from Summary sheet
for row in range(3, 12):
    label = ws.cell(row, 1).value
    value = ws.cell(row, 2).value
    if label and value is not None:
        print(f"{label:30} {value}")

print()
print("=" * 100)

# Check if there's a Mismatches sheet
if 'Mismatches' in wb.sheetnames:
    ws_mm = wb['Mismatches']
    mismatch_count = ws_mm.max_row - 1  # Subtract header row

    print(f"MISMATCHES SHEET: {mismatch_count} mismatches found")
    print("=" * 100)
    print()

    # Print header
    headers = []
    for col in range(1, 7):
        headers.append(str(ws_mm.cell(1, col).value or ""))
    print(" | ".join(f"{h:20}" for h in headers))
    print("-" * 140)

    # Print first 10 mismatches
    for row in range(2, min(12, ws_mm.max_row + 1)):
        values = []
        for col in range(1, 7):
            val = ws_mm.cell(row, col).value
            values.append(str(val or "")[:20])
        print(" | ".join(f"{v:20}" for v in values))

    if mismatch_count > 10:
        print(f"\n... and {mismatch_count - 10} more mismatches (see Excel file for full details)")
else:
    print("✅ NO MISMATCHES FOUND - PERFECT MATCH!")
    print("=" * 100)

print()
print(f"📄 Full report available at: {report_path}")
print("=" * 100)
