"""
Map all Section 3 columns in Excel to understand the complete structure
"""

import openpyxl

excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"

print("="*100)
print("SECTION 3 - COMPLETE COLUMN MAPPING")
print("="*100)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['EMEI']

# Section 3 data starts at row 77, column 3 is day number
day1_row = 77

print("\nDay 1 data (Row 77) - showing all columns:")
for col_idx in range(1, 25):  # Check columns A-X
    cell_val = ws.cell(day1_row, col_idx).value
    if cell_val is not None and cell_val != "":
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): {cell_val}")

# Check day 9 (should have observation "DIA DA FAMÍLIA NA ESCOLA")
day9_row = 85  # 77 + 8
print(f"\nDay 9 data (Row {day9_row}) - showing all columns:")
for col_idx in range(1, 25):
    cell_val = ws.cell(day9_row, col_idx).value
    if cell_val is not None and cell_val != "":
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): {cell_val}")

# Check Total row (should be row 77 + 31 = 108)
total_row = 108
print(f"\nTotal row (Row {total_row}) - showing all columns:")
for col_idx in range(1, 25):
    cell_val = ws.cell(total_row, col_idx).value
    if cell_val is not None and cell_val != "":
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): {cell_val}")

# Now let's verify the column headers
header_row = 74
print(f"\nColumn headers (Row {header_row}):")
for col_idx in range(1, 25):
    cell_val = ws.cell(header_row, col_idx).value
    if cell_val is not None and cell_val != "":
        print(f"  Column {col_idx} (Excel {chr(64+col_idx)}): {cell_val}")

# Check which column has "Lanche (6h)" for Group B
print(f"\nSearching for Group B 'Lanche (6h)' column...")
for col_idx in range(1, 25):
    header_val = ws.cell(header_row, col_idx).value
    group_header = ws.cell(73, col_idx).value  # Row 73 has Group A/B labels
    if header_val and "Lanche" in str(header_val) and "6h" in str(header_val):
        print(f"  Found at Column {col_idx}: '{header_val}' (Group: {group_header})")
        # Check day 1 value in this column
        day1_val = ws.cell(day1_row, col_idx).value
        print(f"    Day 1 value: {day1_val}")

print("\n" + "="*100)
print("EXPECTED STRUCTURE (from user):")
print("="*100)
print("""
Columns:
1. Dia (day number)
2. Dieta_Especial_Grupo_A_Frequencia
3. Dieta_Especial_Grupo_A_Lanche_4H
4. Dieta_Especial_Grupo_A_Lanche_6H
5. Dieta_Especial_Grupo_A_Refeicao_Dieta_Enteral
6. Dieta_Especial_Grupo_B_Frequencia
7. Dieta_Especial_Grupo_B_Lanche_4H
8. Dieta_Especial_Grupo_B_Lanche_6H
9. Lanche_Emergencial
10. Kit_Lanche
11. Observacoes

Day 1 should have:
  - Grupo_B_Frequencia = 16 (empty for Group A)
  - Grupo_B_Lanche_6H = 16

Day 9 should have:
  - Grupo_B_Frequencia = 4
  - Grupo_B_Lanche_6H = 4
  - Observacoes = "DIA DA FAMÍLIA NA ESCOLA"

Total row:
  - Grupo_B_Frequencia = 332
  - Grupo_B_Lanche_6H = 332
""")