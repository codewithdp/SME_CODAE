"""
Validate Section 3 structure in the Excel file
Based on user's description:
- 11 columns (Dia + 4 Group A fields + 3 Group B fields + 2 emergency fields + Observacoes)
- 31 days + Total row
"""

import openpyxl

excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"

print("="*100)
print("SECTION 3 - VALIDATING EXCEL STRUCTURE")
print("="*100)

# Open workbook
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['EMEI']

print("\nSearching for Section 3 table...")
print("Looking for: 'TABELA 3' or 'INFORMAR A QUANTIDADE DE ALIMENTAÇÃO FORNECIDA POR DIA E POR DIETA ESPECIAL'")

# Scan for Section 3 header
section3_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=False), start=1):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "TABELA 3" in cell.value.upper() or "DIETA ESPECIAL AUTORIZADA" in cell.value.upper():
                section3_row = row_idx
                print(f"\n✅ Found Section 3 at row {row_idx}: {cell.value[:80]}...")
                break
    if section3_row:
        break

if not section3_row:
    print("❌ Section 3 not found!")
    exit(1)

# Check the structure starting from section3_row
print(f"\nExamining structure starting at row {section3_row}...")

# Show next 5 rows to see headers
print(f"\nHeader rows (rows {section3_row} to {section3_row + 4}):")
for i in range(5):
    row_num = section3_row + i
    row_data = [cell.value for cell in ws[row_num]]
    # Show first 12 columns
    print(f"  Row {row_num}: {row_data[:12]}")

# Find where the day numbers start (should be after headers)
day_start_row = None
for i in range(5):
    row_num = section3_row + i
    first_col_val = ws.cell(row_num, 1).value
    if first_col_val == 1 or first_col_val == "1":
        day_start_row = row_num
        print(f"\n✅ Day data starts at row {day_start_row}")
        break

if not day_start_row:
    print("⚠️  Could not find day data start")
else:
    # Show sample days
    print(f"\nSample data rows:")
    for day in [1, 9, 29]:  # Days with observations
        row_num = day_start_row + day - 1
        row_data = [cell.value for cell in ws[row_num]]
        print(f"\n  Day {day} (row {row_num}):")
        print(f"    Col 0 (Dia): {row_data[0]}")
        print(f"    Col 1 (Grupo A Freq): {row_data[1]}")
        print(f"    Col 5 (Grupo B Freq): {row_data[5] if len(row_data) > 5 else 'N/A'}")
        print(f"    Col 7 (Grupo B Lanche 6H): {row_data[7] if len(row_data) > 7 else 'N/A'}")
        print(f"    Col 10 (Observacoes): {row_data[10] if len(row_data) > 10 else 'N/A'}")

    # Check Total row
    total_row_num = day_start_row + 31
    total_row_data = [cell.value for cell in ws[total_row_num]]
    print(f"\n  Total row (row {total_row_num}):")
    print(f"    Col 0: {total_row_data[0]}")
    print(f"    Col 5 (Grupo B Freq): {total_row_data[5] if len(total_row_data) > 5 else 'N/A'}")
    print(f"    Col 7 (Grupo B Lanche 6H): {total_row_data[7] if len(total_row_data) > 7 else 'N/A'}")

print("\n" + "="*100)
print("VALIDATION")
print("="*100)
print("Expected structure:")
print("  - 11 columns: Dia + 4 Group A + 3 Group B + 2 Emergency + Observacoes")
print("  - 32 rows: 31 days + Total")
print("  - Group B Frequencia (col 5) and Lanche_6H (col 7) have matching values")
print("  - Total for Group B: 332")
