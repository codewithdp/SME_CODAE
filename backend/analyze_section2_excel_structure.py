"""
Analyze Section 2 Excel structure to map all 36 columns
"""

import openpyxl

excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["EMEI"]

print("="*100)
print("SECTION 2 - EXCEL COLUMN STRUCTURE ANALYSIS")
print("="*100)

# Section 2 starts at row 28 (day 1) and goes to row 58 (day 31)
# Let's look at the header rows to understand column structure

print("\nHEADER ROWS (to identify column structure):")
print("-"*100)

# Look at rows 23-27 (likely header rows above the data)
for row_num in range(23, 28):
    print(f"\nRow {row_num}:")
    # Check columns A through BZ (first 78 columns should be enough)
    for col_num in range(1, 79):
        cell_value = ws.cell(row_num, col_num).value
        if cell_value:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_num)
            print(f"  {col_letter}{row_num} (col {col_num}): {str(cell_value)[:60]}")

print("\n" + "="*100)
print("DAY 1 DATA (Row 28) - to see which columns have data:")
print("-"*100)

day1_row = 28
print(f"\nRow {day1_row} (Day 1):")
for col_num in range(1, 79):
    cell_value = ws.cell(day1_row, col_num).value
    if cell_value is not None and str(cell_value).strip():
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_num)
        print(f"  {col_letter} (col {col_num}): {cell_value}")

print("\n" + "="*100)
print("STRUCTURE MAPPING:")
print("-"*100)

# Based on earlier analysis, we know some columns:
# Col C (3): Day number
# P1:
#   Col R (18): Frequencia
#   Col U (21): Lanche (6h)
#   Col X (24): Refeicao
#   Col AB (28): Repeticao_refeicao
#   Col AE (31): Sobremesa
#   Col AI (35): Repeticao_sobremesa

# P3:
#   Col AU (47): Frequencia
#   Col AY (51): Lanche (6h)
#   Col BD (57): Refeicao
#   Col BH (61): Repeticao_refeicao
#   Col BL (62): Sobremesa
#   Col BT (69): Repeticao_sobremesa

print("\nKnown P1 columns:")
print("  Col R (18): Frequencia")
print("  Col U (21): Lanche 6h")
print("  Col X (24): Refeicao")
print("  Col AB (28): Repeticao refeicao")
print("  Col AE (31): Sobremesa")
print("  Col AI (35): Repeticao sobremesa")

print("\nKnown P3 columns:")
print("  Col AU (47): Frequencia")
print("  Col AY (51): Lanche 6h")
print("  Col BD (57): Refeicao")
print("  Col BH (61): Repeticao refeicao")
print("  Col BL (62): Sobremesa")
print("  Col BT (69): Repeticao sobremesa")

print("\nNeed to find:")
print("  - INTEGRAL columns (11 fields)")
print("  - P1 Lanche 4h")
print("  - INTERMEDIÁRIO columns (6 fields)")
print("  - P3 Lanche 4h")
print("  - Doce checkboxes (4 columns)")
