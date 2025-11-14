"""
Detailed exploration of Section 3 Excel structure
"""

import openpyxl

excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"

print("="*100)
print("SECTION 3 - DETAILED STRUCTURE")
print("="*100)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['EMEI']

# Section 3 starts at row 72
section3_start = 72

print(f"\nShowing rows {section3_start} to {section3_start + 10}, columns A-P (0-15):")
print()

for row_idx in range(section3_start, section3_start + 11):
    row_data = []
    for col_idx in range(16):  # Columns A-P (0-15)
        cell = ws.cell(row_idx, col_idx + 1)
        val = cell.value
        if val is None:
            row_data.append("")
        elif isinstance(val, str):
            row_data.append(val[:20])  # Truncate long strings
        else:
            row_data.append(str(val)[:20])

    print(f"Row {row_idx:3d}: {' | '.join(f'{v:20s}' for v in row_data)}")

print(f"\n\nShowing more data rows (rows {section3_start + 5} to {section3_start + 15}):")
print()

for row_idx in range(section3_start + 5, section3_start + 16):
    row_data = []
    for col_idx in range(16):
        cell = ws.cell(row_idx, col_idx + 1)
        val = cell.value
        if val is None:
            row_data.append("")
        else:
            row_data.append(str(val)[:15])

    print(f"Row {row_idx:3d}: {' | '.join(f'{v:15s}' for v in row_data)}")

# Try to find where day "1" appears
print(f"\n\nSearching for day '1' in rows {section3_start} to {section3_start + 20}...")
for row_idx in range(section3_start, section3_start + 21):
    for col_idx in range(1, 16):
        cell_val = ws.cell(row_idx, col_idx).value
        if cell_val == 1 or cell_val == "1":
            print(f"  Found '1' at Row {row_idx}, Column {col_idx} (Excel col {chr(64+col_idx)})")
            # Show context
            print(f"    Surrounding values:")
            for c in range(max(1, col_idx-2), min(16, col_idx+3)):
                print(f"      Col {c}: {ws.cell(row_idx, c).value}")
