"""
Verify Section 1 mapping for both Excel and PDF
"""

import os
from dotenv import load_dotenv
load_dotenv()

print("="*80)
print("SECTION 1 MAPPING VERIFICATION")
print("="*80)

# ============================================================================
# EXCEL MAPPING
# ============================================================================
print("\n1. EXCEL MAPPING:")
print("-" * 80)

import openpyxl
excel_path = "/Users/writetodennis/dev/SME/019382.xlsm"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["EMEI"]

print("\nExpected structure from excel_parser_custom.py:")
print("  Row 15: INTEGRAL")
print("  Row 16: 1º PERÍODO MATUTINO")
print("  Row 17: 2º PERÍODO INTERMEDIÁRIO")
print("  Row 18: 3º PERÍODO VESPERTINO")
print("  Row 20: Total")
print("\n  Col F (6): Period names")
print("  Col L (12): Number of students")
print("  Col R (18): Special Diet A")
print("  Col V (22): Special Diet B")

print("\n" + "-" * 80)
print("ACTUAL EXCEL DATA:")
print("-" * 80)

rows_to_check = [15, 16, 17, 18, 20]
for row_num in rows_to_check:
    period_name = ws.cell(row_num, 6).value  # Col F
    students = ws.cell(row_num, 12).value     # Col L
    diet_a = ws.cell(row_num, 18).value       # Col R
    diet_b = ws.cell(row_num, 22).value       # Col V

    print(f"Row {row_num}:")
    print(f"  Period (F): {period_name}")
    print(f"  Students (L): {students}")
    print(f"  Diet A (R): {diet_a}")
    print(f"  Diet B (V): {diet_b}")
    print()

# ============================================================================
# PDF MAPPING
# ============================================================================
print("\n" + "="*80)
print("2. PDF MAPPING:")
print("-" * 80)

from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential

endpoint = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
key = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")

client = DocumentIntelligenceClient(endpoint, AzureKeyCredential(key))

pdf_path = "/Users/writetodennis/dev/SME/EMEI_test1.pdf"

with open(pdf_path, "rb") as f:
    poller = client.begin_analyze_document(
        model_id="prebuilt-layout",
        analyze_request=f.read(),
        content_type="application/octet-stream"
    )
    result = poller.result()

# Table 2 is Section 1 (index 1)
table = result.tables[1]

print(f"\nTable 2 (Section 1): {table.row_count} rows × {table.column_count} columns")

# Build 2D array
cells = [[None for _ in range(table.column_count)] for _ in range(table.row_count)]

for cell in table.cells:
    row = cell.row_index
    col = cell.column_index
    row_span = getattr(cell, 'row_span', 1) or 1
    col_span = getattr(cell, 'column_span', 1) or 1

    for r in range(row, min(row + row_span, table.row_count)):
        for c in range(col, min(col + col_span, table.column_count)):
            if r < len(cells) and c < len(cells[0]):
                cells[r][c] = cell.content

print("\nExpected PDF structure:")
print("  Col 0: Period names")
print("  Col 1: Hours of attendance")
print("  Col 2: Number of students")
print("  Col 3: Special Diet A")
print("  Col 4: Special Diet B")

print("\n" + "-" * 80)
print("ACTUAL PDF DATA (Table 2):")
print("-" * 80)

for row_idx in range(table.row_count):
    print(f"\nRow {row_idx}:")
    for col_idx in range(table.column_count):
        content = cells[row_idx][col_idx]
        print(f"  Col {col_idx}: {content}")

# ============================================================================
# COMPARISON
# ============================================================================
print("\n" + "="*80)
print("3. COMPARISON WITH USER PROVIDED DATA:")
print("-" * 80)

print("\nUser's table:")
print("  Integral (8 horas): — students, — Diet A, — Diet B")
print("  1º Período Matutino: 294 students, — Diet A, 12 Diet B")
print("  2º Período Intermediário: — students, — Diet A, — Diet B")
print("  3º Período Vespertino: 296 students, — Diet A, 6 Diet B")
print("  Total: 590 students, 0 Diet A, 18 Diet B")

print("\nExcel extracted:")
excel_data = {
    15: ("INTEGRAL", ws.cell(15, 12).value, ws.cell(15, 18).value, ws.cell(15, 22).value),
    16: ("1º PERÍODO MATUTINO", ws.cell(16, 12).value, ws.cell(16, 18).value, ws.cell(16, 22).value),
    17: ("2º PERÍODO INTERMEDIÁRIO", ws.cell(17, 12).value, ws.cell(17, 18).value, ws.cell(17, 22).value),
    18: ("3º PERÍODO VESPERTINO", ws.cell(18, 12).value, ws.cell(18, 18).value, ws.cell(18, 22).value),
    20: ("TOTAL", ws.cell(20, 12).value, ws.cell(20, 18).value, ws.cell(20, 22).value),
}

for row, (name, students, diet_a, diet_b) in excel_data.items():
    print(f"  Row {row} - {name}: {students} students, {diet_a} Diet A, {diet_b} Diet B")
