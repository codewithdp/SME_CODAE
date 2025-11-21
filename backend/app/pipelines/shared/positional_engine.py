"""
Shared Positional Reconciliation Engine Logic
"""
import os
import logging
from typing import Dict, List, Optional, Tuple, Any
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import load_workbook
import xlrd
from dotenv import load_dotenv

from ...pdf_cell_image_extractor import PDFCellImageExtractor

load_dotenv()

logger = logging.getLogger(__name__)


class XlrdSheetWrapper:
    """Wrapper for xlrd sheet to provide openpyxl-like interface"""

    def __init__(self, sheet):
        self.sheet = sheet

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        """Iterate over rows, mimicking openpyxl interface"""
        if max_row is None:
            max_row = self.sheet.nrows

        for row_idx in range(min_row - 1, max_row):  # Convert to 0-indexed
            if values_only:
                yield tuple(self.sheet.cell_value(row_idx, col) for col in range(self.sheet.ncols))
            else:
                yield tuple(self.sheet.cell(row_idx, col) for col in range(self.sheet.ncols))


class PositionalReconciliationEngine:
    """Reconciliation engine using fixed positional column mapping"""

    def __init__(self):
        self.endpoint = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
        self.key = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")

        # Determine which model to use based on configuration
        model_type = os.getenv("AZURE_DI_MODEL_TYPE", "prebuilt").lower()
        if model_type == "custom":
            self.model_id = os.getenv("CUSTOM_MODEL_NAME", "SEM_EMEI_v1")
            logger.info(f"Using custom Azure DI model: {self.model_id}")
        else:
            self.model_id = "prebuilt-layout"
            logger.info("Using prebuilt-layout Azure DI model")

        if not self.endpoint or not self.key:
            raise ValueError("Azure Document Intelligence credentials not found in environment")

        self.client = DocumentIntelligenceClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.key)
        )

        # Initialize PDF cell image extractor for mismatch visualization
        self.image_extractor = PDFCellImageExtractor(zoom_factor=1.0)
        self.extract_images = False  # DISABLED: Set to False to disable image extraction for faster processing

    @staticmethod
    def excel_column_letter(col_idx: int) -> str:
        """Convert column index (0-based) to Excel column letter"""
        result = ""
        col_idx += 1  # Convert to 1-based
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + ord('A')) + result
            col_idx //= 26
        return result

    def extract_table(self, pdf_path: str, table_index: int = 2) -> Tuple[Any, Dict]:
        """
        Extract a table from PDF using Azure DI
        Args:
            pdf_path: Path to PDF file
            table_index: Index of table to extract (default 2 for Section2)
        Returns:
            (table_object, metadata)
        """
        with open(pdf_path, "rb") as f:
            poller = self.client.begin_analyze_document(
                self.model_id,
                analyze_request=f,
                content_type="application/pdf"
            )
            result = poller.result()

        if not result.tables or len(result.tables) <= table_index:
            raise ValueError(f"Table {table_index} not found in PDF")

        table = result.tables[table_index]

        metadata = {
            "row_count": table.row_count,
            "column_count": table.column_count,
            "table_index": table_index
        }

        return table, metadata

    def find_day_1_row(self, table: Any, search_start_row: int = 1) -> int:
        """
        Find the row index where Day 1 actually starts in the PDF table

        Args:
            table: PDF table object
            search_start_row: Row to start searching from (default 1, after title row)

        Returns:
            Row index where day=1 is found, or search_start_row if not found
        """
        for row_idx in range(search_start_row, min(search_start_row + 5, table.row_count)):
            # Get day column value (column 0)
            day_val = ""
            for cell in table.cells:
                if cell.row_index == row_idx and cell.column_index == 0:
                    day_val = (cell.content or "").strip()
                    break

            # Check if this row has day=1
            if day_val == "1" or day_val == 1:
                logger.info(f"Found Day 1 at PDF row {row_idx}")
                return row_idx

        # Default to search_start_row + 1 if not found
        logger.warning(f"Day 1 not found, defaulting to row {search_start_row + 1}")
        return search_start_row + 1

    def build_pdf_table_structure(self, table: Any, data_start_row: int = 2, auto_detect_day1: bool = True) -> Dict:
        """
        Build a structured representation of the PDF table
        Args:
            table: PDF table object
            data_start_row: Row index where data starts (default 2 for Section2/3, 1 for Section1)
            auto_detect_day1: If True, automatically find Day 1 row. If False, use data_start_row as-is.
                              Set to False for Section 1 which doesn't have day numbers.
        Returns: {
            "rows": [
                {"row_idx": idx, "day": day_num, "cells": {col_idx: content}},
                ..
            ],
            "actual_data_start_row": int  # Where Day 1 actually is
        }
        """
        # Find where Day 1 actually starts (handles different header row counts)
        # Only do this for Section 2 and 3 which have day-based rows
        if auto_detect_day1:
            actual_data_start = self.find_day_1_row(table, search_start_row=data_start_row - 1)
        else:
            actual_data_start = data_start_row

        # Extract data rows (skip header rows)
        rows = []
        for row_idx in range(actual_data_start, table.row_count):
            row_data = {
                "row_idx": row_idx,
                "cells": {},
                "cell_objects": {}  # Store original Azure DI cell objects for image extraction
            }

            # Get all cells for this row
            for cell in table.cells:
                if cell.row_index == row_idx:
                    row_data["cells"][cell.column_index] = cell.content if cell.content else ""
                    row_data["cell_objects"][cell.column_index] = cell  # Store cell object

            # Get day number from column 0
            day_val = row_data["cells"].get(0, "").strip()
            if day_val and day_val.lower() != "total":
                try:
                    row_data["day"] = int(day_val)
                except (ValueError, AttributeError):
                    row_data["day"] = day_val
            elif "total" in day_val.lower():
                row_data["day"] = "Total"
            else:
                row_data["day"] = None

            rows.append(row_data)

        return {
            "rows": rows,
            "actual_data_start_row": actual_data_start
        }

    def load_excel_sheet(self, excel_path: str, sheet_name: str = "EMEI") -> Any:
        """Load Excel sheet - supports both .xlsx and .xls formats"""

        # Check file extension to determine which library to use
        if excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx'):
            # Use xlrd for old .xls format
            wb = xlrd.open_workbook(excel_path)
            ws = None

            for sn in wb.sheet_names():
                if sheet_name.upper() in sn.upper():
                    ws = wb.sheet_by_name(sn)
                    break

            if not ws:
                raise ValueError(f"Sheet containing '{sheet_name}' not found in Excel file")

            # Wrap xlrd sheet to provide similar interface to openpyxl
            return XlrdSheetWrapper(ws)
        else:
            # Use openpyxl for .xlsx format
            wb = load_workbook(excel_path, data_only=True)
            ws = None

            for sn in wb.sheetnames:
                if sheet_name.upper() in sn.upper():
                    ws = wb[sn]
                    break

            if not ws:
                raise ValueError(f"Sheet containing '{sheet_name}' not found in Excel file")

            return ws

    @staticmethod
    def _normalize_value(value: str) -> str:
        """
        Normalize value for comparison

        Handles:
        1. Empty cells vs :unselected: (both mean unchecked/empty)
        2. Checkbox representations (X/x vs :selected:)
        3. Number formatting (remove thousands separators)
        4. Values with checkbox markers (e.g., ": 0 :unselected:" → "0")
        """
        if not value or value == "":
            return ""

        # Clean up whitespace and newlines first
        value = value.replace("\n", " ").strip()

        # Treat :unselected: as empty (unchecked checkbox = empty cell)
        if value == ":unselected:" or value == ": unselected :" or value == ":unselected :":
            return ""

        # Handle values with :unselected: prefix/suffix (e.g., ": 0 :unselected:" → "0")
        if ":unselected:" in value or "unselected" in value.lower():
            cleaned = value.replace(":unselected:", "").replace(": unselected :", "")
            cleaned = cleaned.replace(":", "").strip()
            # If the cleaned value is empty, treat as empty cell
            if not cleaned:
                return ""
            # Otherwise return the cleaned value (e.g., "0")
            value = cleaned

        # Normalize checkbox values
        if value.upper().strip() == "X":
            return ""  # Treat X as empty for CEI

        # Handle values that contain both content and checkbox markers
        # e.g., "0 :selected:", "x :selected:", "8 :selected:", "4\n:selected:"
        if ":selected:" in value or "selected" in value.lower():
            # Strip the :selected: marker and any extra colons/spaces/newlines
            cleaned = value.replace(":selected:", "").replace(": selected :", "")
            cleaned = cleaned.replace(":", "").strip()
            # If it's just "x" or "X", treat as empty
            if cleaned.upper() == "X" or cleaned == "":
                return ""
            # Otherwise return the cleaned value without the checkbox marker
            # e.g., "8 :selected:" → "8", "4\n:selected:" → "4"
            value = cleaned

        # Number normalization
        # Handle trailing .0 (e.g., "2.0" → "2")
        if value.endswith(".0"):
            value = value[:-2]

        # Treat "0" and "-" as empty for reconciliation
        if value == "0" or value == "-":
            return ""

        # AGGRESSIVE number normalization: Remove ALL dots and commas from numbers
        # This handles: 1.665 → 1665, 4,667 → 4667, 1.234.567 → 1234567, etc.
        # Works for all sections, not just totals
        if "." in value or "," in value:
            # Try removing all dots and commas
            cleaned_number = value.replace(".", "").replace(",", "")
            # If the result is all digits, return the cleaned version
            if cleaned_number.isdigit():
                return cleaned_number
            # Otherwise keep the original value (might be a decimal like 1.5)

        return value

    def extract_pdf_headers(self, table: Any, header_rows: int = 2) -> Dict[int, str]:
        """
        Extract column headers from PDF table

        Args:
            table: PDF table object
            header_rows: Number of header rows to check (default 2)

        Returns:
            Dict mapping column index to header text
        """
        headers = {}

        for cell in table.cells:
            if cell.row_index < header_rows:
                col_idx = cell.column_index
                content = (cell.content or "").strip()

                # Combine headers from multiple rows if they exist
                if col_idx in headers:
                    if content and content not in headers[col_idx]:
                        headers[col_idx] = f"{headers[col_idx]} {content}".strip()
                else:
                    headers[col_idx] = content

        return headers

    @staticmethod
    def normalize_header(text: str) -> str:
        """Normalize header text for comparison"""
        if not text:
            return ""
        # Lowercase, remove extra whitespace, normalize common variations
        normalized = text.lower().strip()

        # Remove checkbox markers and data that gets mixed into headers
        normalized = normalized.replace(":selected:", "").replace(":unselected:", "")

        # Remove standalone numbers that are likely data values (not part of age groups)
        # Keep numbers that are part of patterns like "01 a 03" but remove standalone "2"
        import re
        # Remove numbers at end of string (likely data values)
        normalized = re.sub(r'\s+\d+$', '', normalized)
        # Remove numbers at start that aren't part of age patterns
        normalized = re.sub(r'^\d+\s+(?![a])', '', normalized)

        normalized = " ".join(normalized.split())  # Collapse whitespace

        # Remove common accent variations
        normalized = normalized.replace("á", "a").replace("é", "e").replace("ã", "a")
        normalized = normalized.replace("ç", "c").replace("ô", "o").replace("ê", "e")
        return normalized

    def calculate_header_similarity(self, header1: str, header2: str) -> float:
        """
        Calculate similarity between two headers
        Returns value between 0 and 1
        Handles OCR word reordering (e.g., "01 a 03 anos e 11 meses" vs "e 11 meses 01 a 03 anos")
        """
        h1 = self.normalize_header(header1)
        h2 = self.normalize_header(header2)

        if not h1 or not h2:
            return 0.0

        if h1 == h2:
            return 1.0

        # Check word set equality (handles reordered words from OCR)
        words1 = set(h1.split())
        words2 = set(h2.split())

        if words1 == words2:
            return 1.0  # Same words, different order = perfect match

        # Check if one contains the other
        if h1 in h2 or h2 in h1:
            return 0.8

        # Check word overlap ratio
        if words1 and words2:
            overlap = len(words1 & words2)
            total = len(words1 | words2)
            similarity = overlap / total if total > 0 else 0.0
            # Boost score if most words match
            if similarity >= 0.8:
                return 0.9
            return similarity

        return 0.0

    def build_dynamic_column_mapping(
        self,
        pdf_table: Any,
        excel_column_names: Dict[int, str],
        base_mapping: Dict[int, int],
        header_rows: int = 2
    ) -> Dict[int, int]:
        """
        Build dynamic column mapping by matching PDF headers to Excel headers

        Args:
            pdf_table: PDF table object
            excel_column_names: Dict mapping Excel col idx to header name
            base_mapping: Original positional mapping to use as fallback
            header_rows: Number of PDF header rows

        Returns:
            Adjusted column mapping
        """
        pdf_headers = self.extract_pdf_headers(pdf_table, header_rows)

        if not pdf_headers:
            logger.warning("No PDF headers found, using base mapping")
            return base_mapping

        logger.info(f"Building dynamic mapping from {len(excel_column_names)} Excel headers to {len(pdf_headers)} PDF columns")

        dynamic_mapping = {}
        unmatched_excel = []

        for excel_col, excel_header in excel_column_names.items():
            if not excel_header:  # Skip empty headers
                # Use base mapping for empty headers
                if excel_col in base_mapping:
                    dynamic_mapping[excel_col] = base_mapping[excel_col]
                continue

            # Find best matching PDF column
            best_match_col = None
            best_match_score = 0.0

            for pdf_col, pdf_header in pdf_headers.items():
                score = self.calculate_header_similarity(excel_header, pdf_header)
                if score > best_match_score:
                    best_match_score = score
                    best_match_col = pdf_col

            if best_match_score >= 0.5:  # Minimum threshold for match
                dynamic_mapping[excel_col] = best_match_col
                logger.debug(f"Matched Excel col {excel_col} '{excel_header}' -> PDF col {best_match_col} (score: {best_match_score:.2f})")
            else:
                # Fall back to base mapping
                if excel_col in base_mapping:
                    dynamic_mapping[excel_col] = base_mapping[excel_col]
                    logger.debug(f"No match for '{excel_header}', using base mapping: {excel_col} -> {base_mapping[excel_col]}")
                else:
                    unmatched_excel.append(excel_col)

        if unmatched_excel:
            logger.warning(f"Unmatched Excel columns: {unmatched_excel}")

        logger.info(f"Dynamic mapping complete: {len(dynamic_mapping)} columns mapped")
        return dynamic_mapping

    def reconcile_section(
        self,
        pdf_path: str,
        excel_path: str,
        table_index: int = 2,
        excel_start_row: int = 28,
        column_mapping: Dict[int, int] = None,
        column_names: Dict[int, str] = None,
        pdf_data_start_row: int = 2,
        excel_row_skip: int = 0,
        pdf_table: Any = None,  # OPTIMIZATION: Pass pre-extracted table to avoid re-analyzing PDF
        sheet_name: str = "EMEI",  # Sheet name pattern to search for
        use_dynamic_mapping: bool = False,  # Enable dynamic header-based mapping
        dynamic_header_rows: int = None,  # Explicit header rows for dynamic mapping
        auto_detect_day1: bool = None  # Auto-detect Day 1 row (None=auto based on data_start)
    ) -> Dict:
        """
        Reconcile a section between PDF table and Excel using positional mapping

        Args:
            pdf_path: Path to PDF file
            excel_path: Path to Excel file
            table_index: Index of table in PDF (0 for Section1, 2 for Section2, 4 for Section3)
            excel_start_row: Starting row in Excel (1-indexed) for first data day
            column_mapping: Dict mapping Excel col idx to PDF col idx
            pdf_data_start_row: PDF row index where Day 1 data starts (1 for Section1, 2 for Section2, 3 for Section3)
            excel_row_skip: Extra offset for Total row (1 for Section1 due to empty row 19)
            pdf_table: Optional pre-extracted PDF table object (avoids re-analyzing PDF)
            sheet_name: Sheet name pattern to search for in Excel file (default "EMEI")

        Returns:
            Dictionary with reconciliation results
        """
        if column_mapping is None:
            # Fallback if no mapping provided (should not happen in new architecture)
            return {"error": "No column mapping provided"}

        # Extract PDF table (or use provided one)
        if pdf_table is not None:
            # Use pre-extracted table (optimization to avoid redundant Azure DI API calls)
            pdf_metadata = {
                "row_count": pdf_table.row_count,
                "column_count": pdf_table.column_count,
                "table_index": table_index
            }
            logger.info(f"Using pre-extracted PDF table (avoiding redundant API call)")
        else:
            # Extract table from PDF (backward compatibility)
            pdf_table, pdf_metadata = self.extract_table(pdf_path, table_index)
        # Determine auto_detect_day1: use explicit param if set, otherwise heuristic
        should_detect_day1 = auto_detect_day1 if auto_detect_day1 is not None else (pdf_data_start_row >= 2)
        pdf_structure = self.build_pdf_table_structure(
            pdf_table,
            data_start_row=pdf_data_start_row,
            auto_detect_day1=should_detect_day1
        )

        # Use the actual detected Day 1 row for calculations
        actual_pdf_data_start = pdf_structure.get("actual_data_start_row", pdf_data_start_row)
        logger.info(f"Day 1 detected at PDF row {actual_pdf_data_start} (expected around row {pdf_data_start_row})")

        # Load Excel
        ws = self.load_excel_sheet(excel_path, sheet_name=sheet_name)

        # Use dynamic mapping if explicitly enabled and column_names provided
        # This handles OCR variations like word reordering in headers
        if use_dynamic_mapping and column_names:
            # Use explicit header_rows if provided, otherwise use pdf_data_start_row
            header_rows = dynamic_header_rows if dynamic_header_rows else pdf_data_start_row
            column_mapping = self.build_dynamic_column_mapping(
                pdf_table,
                column_names,
                column_mapping,
                header_rows=header_rows
            )
            logger.info(f"Using dynamic column mapping: {len(column_mapping)} columns mapped (header_rows={header_rows})")
        else:
            logger.info(f"Using fixed column mapping: {len(column_mapping)} columns mapped")

        # Reconciliation results
        results = {
            "table_index": table_index,
            "pdf_metadata": pdf_metadata,
            "excel_start_row": excel_start_row,
            "column_mapping": {f"Excel_{k}": f"PDF_{v}" for k, v in column_mapping.items()},
            "days_compared": 0,
            "cells_compared": 0,
            "matches": 0,
            "mismatches": 0,
            "match_percentage": 0.0,
            "day_results": [],
            "mismatched_days": []
        }

        # Process each PDF row
        for pdf_row in pdf_structure["rows"]:
            if pdf_row["day"] is None:
                continue

            day_num = pdf_row["day"]

            # Check if this is the Total row - if so, process it and then stop
            is_total_row = isinstance(day_num, str) and day_num.lower() == "total"

            # Calculate Excel row for this day
            if isinstance(day_num, int):
                excel_row_idx = excel_start_row + (day_num - 1)
            else:
                # For Total or other special rows, use positional matching
                # Use actual_pdf_data_start (detected Day 1 row) instead of pdf_data_start_row parameter
                excel_row_idx = excel_start_row + (pdf_row["row_idx"] - actual_pdf_data_start)
                # Add excel_row_skip ONLY for Total row (handles empty rows in Excel like Section1 row 19)
                if isinstance(day_num, str) and day_num.lower() == "total":
                    excel_row_idx += excel_row_skip

            excel_row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]

            # Compare cells using column mapping
            day_matches = 0
            day_mismatches = 0
            day_cells_compared = 0
            mismatched_cells = []

            for excel_col_idx, pdf_col_idx in column_mapping.items():
                # Get Excel value
                excel_value = excel_row[excel_col_idx] if excel_col_idx < len(excel_row) else None
                excel_str = str(excel_value).strip() if excel_value is not None else ""

                # Get PDF value
                pdf_value = pdf_row["cells"].get(pdf_col_idx, "")
                pdf_str = str(pdf_value).strip() if pdf_value else ""

                day_cells_compared += 1

                # Normalize values for comparison
                excel_normalized = self._normalize_value(excel_str)
                pdf_normalized = self._normalize_value(pdf_str)

                # Compare values
                if excel_normalized == pdf_normalized:
                    day_matches += 1
                else:
                    day_mismatches += 1
                    col_letter = self.excel_column_letter(excel_col_idx)
                    excel_cell_ref = f"{col_letter}{excel_row_idx}"

                    # Extract PDF cell image for mismatch visualization (if enabled)
                    pdf_image_base64 = None
                    if self.extract_images:
                        try:
                            cell_obj = pdf_row.get("cell_objects", {}).get(pdf_col_idx)
                            if cell_obj:
                                pdf_image_base64 = self.image_extractor.extract_cell_image_from_azure_cell(
                                    pdf_path=pdf_path,
                                    cell=cell_obj
                                )
                        except Exception as e:
                            logger.warning(f"Failed to extract PDF cell image: {e}")

                    # Get column name if available
                    col_name = None
                    if column_names:
                        col_name = column_names.get(excel_col_idx)

                    # Display normalized values (convert empty to "(empty)" for clarity)
                    def display_value(val):
                        if val == "":
                            return "(empty)"
                        return val

                    mismatched_cells.append({
                        "excel_column": excel_col_idx,
                        "excel_cell_ref": excel_cell_ref,
                        "excel_value": display_value(excel_normalized),
                        "pdf_column": pdf_col_idx,
                        "pdf_value": display_value(pdf_normalized),
                        "excel_row": excel_row_idx,
                        "pdf_image_base64": pdf_image_base64,
                        "column_name": col_name
                    })

            results["days_compared"] += 1
            results["cells_compared"] += day_cells_compared
            results["matches"] += day_matches
            results["mismatches"] += day_mismatches

            match_pct = (day_matches / day_cells_compared * 100) if day_cells_compared > 0 else 0
            row_label = f"Day {day_num}" if isinstance(day_num, int) else str(day_num)

            day_result = {
                "day": str(day_num),
                "label": row_label,
                "excel_row": excel_row_idx,
                "cells_compared": day_cells_compared,
                "matches": day_matches,
                "mismatches": day_mismatches,
                "match_percentage": round(match_pct, 2),
                "status": "matched" if day_mismatches == 0 else "mismatched"
            }

            if mismatched_cells:
                day_result["mismatched_cells"] = mismatched_cells

            results["day_results"].append(day_result)

            if day_mismatches > 0:
                results["mismatched_days"].append(str(day_num))

            # Stop processing after Total row - any rows after Total are not part of the data
            if is_total_row:
                logger.info(f"Reached Total row, stopping processing (processed {results['days_compared']} days)")
                break

        # Calculate overall match percentage
        if results["cells_compared"] > 0:
            results["match_percentage"] = round(
                (results["matches"] / results["cells_compared"]) * 100, 2
            )

        return results
