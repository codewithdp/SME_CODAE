"""
Custom Excel Parser for EMEI 019382 Format
Maps exact cell locations based on the specific Excel template
"""

from typing import List, Optional, Dict, Any
from datetime import datetime
import openpyxl
from pydantic import BaseModel, Field


# ============================================================================
# DATA MODELS
# ============================================================================

class HeaderData(BaseModel):
    """Header section data from Excel"""
    emei_code: str
    emei_name: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    company_name: Optional[str] = None
    month: Optional[str] = None
    year: Optional[str] = None


class EnrollmentPeriod(BaseModel):
    """Student enrollment for a specific period"""
    period_name: str
    num_students: Optional[int] = None
    special_diet_a: Optional[int] = None
    special_diet_b: Optional[int] = None


class Section1Data(BaseModel):
    """Section 1: Student enrollment numbers"""
    periods: List[EnrollmentPeriod]
    total_students: int
    total_special_diet_a: int
    total_special_diet_b: int


class DailyFrequencyIntegral(BaseModel):
    """Daily frequency for INTEGRAL period (11 fields)"""
    day: int
    frequencia: Optional[int] = None
    lanche_4h: Optional[int] = None
    lanche_6h: Optional[int] = None
    refeicao: Optional[int] = None
    repeticao_refeicao: Optional[int] = None
    sobremesa: Optional[int] = None
    repeticao_sobremesa: Optional[int] = None
    refeicao_2a: Optional[int] = None
    repeticao_refeicao_2a: Optional[int] = None
    sobremesa_2a: Optional[int] = None
    repeticao_sobremesa_2a: Optional[int] = None


class DailyFrequencyRecord(BaseModel):
    """Daily frequency for P1/P3 periods (7 fields)"""
    day: int
    frequencia: Optional[int] = None
    lanche_4h: Optional[int] = None
    lanche_6h: Optional[int] = None
    refeicao: Optional[int] = None
    repeticao_refeicao: Optional[int] = None
    sobremesa: Optional[int] = None
    repeticao_sobremesa: Optional[int] = None


class DailyFrequencyIntermediario(BaseModel):
    """Daily frequency for INTERMEDIÁRIO period (6 fields - no Lanche 6h)"""
    day: int
    frequencia: Optional[int] = None
    lanche_4h: Optional[int] = None
    refeicao: Optional[int] = None
    repeticao_refeicao: Optional[int] = None
    sobremesa: Optional[int] = None
    repeticao_sobremesa: Optional[int] = None


class DailyDoceCheckboxes(BaseModel):
    """Daily 'Sobremesa foi doce?' checkboxes (4 fields)"""
    day: int
    integral: Optional[bool] = None
    primeiro_periodo: Optional[bool] = None
    intermediario: Optional[bool] = None
    terceiro_periodo: Optional[bool] = None


class Section2TotalRow(BaseModel):
    """Section 2 TOTAL row data"""
    # INTEGRAL (11 fields)
    integral_frequencia: Optional[int] = None
    integral_lanche_4h: Optional[int] = None
    integral_lanche_6h: Optional[int] = None
    integral_refeicao: Optional[int] = None
    integral_repeticao_refeicao: Optional[int] = None
    integral_sobremesa: Optional[int] = None
    integral_repeticao_sobremesa: Optional[int] = None
    integral_refeicao_2a: Optional[int] = None
    integral_repeticao_refeicao_2a: Optional[int] = None
    integral_sobremesa_2a: Optional[int] = None
    integral_repeticao_sobremesa_2a: Optional[int] = None

    # P1 (7 fields)
    p1_frequencia: Optional[int] = None
    p1_lanche_4h: Optional[int] = None
    p1_lanche_6h: Optional[int] = None
    p1_refeicao: Optional[int] = None
    p1_repeticao_refeicao: Optional[int] = None
    p1_sobremesa: Optional[int] = None
    p1_repeticao_sobremesa: Optional[int] = None

    # INTERMEDIÁRIO (6 fields)
    intermediario_frequencia: Optional[int] = None
    intermediario_lanche_4h: Optional[int] = None
    intermediario_refeicao: Optional[int] = None
    intermediario_repeticao_refeicao: Optional[int] = None
    intermediario_sobremesa: Optional[int] = None
    intermediario_repeticao_sobremesa: Optional[int] = None

    # P3 (7 fields)
    p3_frequencia: Optional[int] = None
    p3_lanche_4h: Optional[int] = None
    p3_lanche_6h: Optional[int] = None
    p3_refeicao: Optional[int] = None
    p3_repeticao_refeicao: Optional[int] = None
    p3_sobremesa: Optional[int] = None
    p3_repeticao_sobremesa: Optional[int] = None


class Section2Data(BaseModel):
    """Section 2: Daily frequency data - ALL periods"""
    integral: List[DailyFrequencyIntegral]
    primeiro_periodo: List[DailyFrequencyRecord]
    intermediario: List[DailyFrequencyIntermediario]
    terceiro_periodo: List[DailyFrequencyRecord]
    doce_checkboxes: List[DailyDoceCheckboxes]
    total: Section2TotalRow


class Section3DayData(BaseModel):
    """Section 3: Daily special diet data (11 fields)"""
    day: int
    # Group A (4 fields)
    grupo_a_frequencia: Optional[int] = None
    grupo_a_lanche_4h: Optional[int] = None
    grupo_a_lanche_6h: Optional[int] = None
    grupo_a_refeicao_enteral: Optional[int] = None
    # Group B (3 fields)
    grupo_b_frequencia: Optional[int] = None
    grupo_b_lanche_4h: Optional[int] = None
    grupo_b_lanche_6h: Optional[int] = None
    # Emergency snacks (2 fields)
    lanche_emergencial: Optional[int] = None
    kit_lanche: Optional[int] = None
    # Observations (1 field)
    observacoes: Optional[str] = None


class Section3TotalRow(BaseModel):
    """Section 3 TOTAL row data"""
    grupo_a_frequencia: Optional[int] = None
    grupo_a_lanche_4h: Optional[int] = None
    grupo_a_lanche_6h: Optional[int] = None
    grupo_a_refeicao_enteral: Optional[int] = None
    grupo_b_frequencia: Optional[int] = None
    grupo_b_lanche_4h: Optional[int] = None
    grupo_b_lanche_6h: Optional[int] = None
    lanche_emergencial: Optional[int] = None
    kit_lanche: Optional[int] = None


class Section3Data(BaseModel):
    """Section 3: Daily special diet data for all 31 days"""
    days: List[Section3DayData]
    total: Section3TotalRow


class ExcelReconciliationData(BaseModel):
    """Complete Excel data"""
    filename: str
    extracted_at: datetime = Field(default_factory=datetime.now)
    header: HeaderData
    section1: Section1Data
    section2: Section2Data
    section3: Optional[Section3Data] = None


# ============================================================================
# EXCEL PARSER CLASS
# ============================================================================

class CustomExcelParser:
    """
    Custom parser for EMEI 019382 Excel format
    Cell mappings based on actual file structure
    """

    def __init__(self):
        # Section 1 row mappings (enrollment)
        self.enrollment_rows = {
            "INTEGRAL": 15,
            "1º PERÍODO MATUTINO": 16,
            "2º PERÍODO INTERMEDIÁRIO": 17,
            "3º PERÍODO VESPERTINO": 18,
        }
        self.enrollment_total_row = 20

        # Section 2 row mappings (daily frequency)
        self.frequency_start_row = 28
        self.frequency_end_row = 58  # Days 1-31

    def parse_file(self, excel_path: str) -> ExcelReconciliationData:
        """Main parsing method"""
        # Load workbook with formulas evaluated
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['EMEI']

        # Extract sections
        header = self._extract_header(ws)
        section1 = self._extract_section1(ws)
        section2 = self._extract_section2(ws)
        section3 = self._extract_section3(ws)

        return ExcelReconciliationData(
            filename=excel_path.split('/')[-1],
            header=header,
            section1=section1,
            section2=section2,
            section3=section3
        )

    # Alias for backwards compatibility
    def parse_excel(self, excel_path: str) -> ExcelReconciliationData:
        """Alias for parse_file"""
        return self.parse_file(excel_path)

    def _extract_header(self, ws) -> HeaderData:
        """Extract header information"""
        emei_code = str(ws.cell(6, 4).value or "").strip()  # D6
        emei_name = str(ws.cell(6, 8).value or "").strip()  # H6
        email = str(ws.cell(7, 8).value or "").strip()  # H7

        # Try to get address (multiple cells)
        address = str(ws.cell(8, 4).value or "").strip()  # D8

        # Company name
        company_name = str(ws.cell(11, 4).value or "").strip()  # D11

        return HeaderData(
            emei_code=emei_code,
            emei_name=emei_name,
            email=email,
            address=address,
            company_name=company_name
        )

    def _extract_section1(self, ws) -> Section1Data:
        """
        Extract Section 1: Enrollment data

        Extracts ALL periods (including empty rows) for complete reconciliation
        """
        periods = []

        # Extract all periods in order (including empty ones)
        for period_name, row_num in self.enrollment_rows.items():
            period_from_excel = ws.cell(row_num, 6).value  # Col F
            num_students = ws.cell(row_num, 12).value  # Col L
            special_diet_a = ws.cell(row_num, 18).value  # Col R
            special_diet_b = ws.cell(row_num, 22).value  # Col V

            # Add ALL periods (even if empty) for complete comparison
            periods.append(EnrollmentPeriod(
                period_name=str(period_from_excel or period_name),
                num_students=int(num_students or 0) if num_students is not None else None,
                special_diet_a=int(special_diet_a or 0) if special_diet_a is not None else None,
                special_diet_b=int(special_diet_b or 0) if special_diet_b is not None else None
            ))

        # Extract totals
        total_students = int(ws.cell(self.enrollment_total_row, 12).value or 0)  # L20
        total_special_a = int(ws.cell(self.enrollment_total_row, 18).value or 0)  # R20
        total_special_b = int(ws.cell(self.enrollment_total_row, 22).value or 0)  # V20

        return Section1Data(
            periods=periods,
            total_students=total_students,
            total_special_diet_a=total_special_a,
            total_special_diet_b=total_special_b
        )

    def _extract_section2(self, ws) -> Section2Data:
        """
        Extract Section 2: Daily frequency data - ALL periods and fields

        Rows: 28-58 (days 1-31), 59 (TOTAL)
        Columns: INTEGRAL (11), P1 (7), INTERMEDIÁRIO (6), P3 (7), DOCE (4)
        """
        integral = []
        primeiro_periodo = []
        intermediario = []
        terceiro_periodo = []
        doce_checkboxes = []

        # Extract ALL 31 days (rows 28-58), including empty ones
        for row_num in range(self.frequency_start_row, self.frequency_end_row + 1):
            day = ws.cell(row_num, 3).value  # Col C
            day_int = self._safe_int(day) if day else (row_num - self.frequency_start_row + 1)

            # INTEGRAL (11 fields) - cols E,G,H,J-Q (5,7,8,10-17)
            integral.append(DailyFrequencyIntegral(
                day=day_int,
                frequencia=self._safe_int(ws.cell(row_num, 5).value),
                lanche_4h=self._safe_int(ws.cell(row_num, 7).value),
                lanche_6h=self._safe_int(ws.cell(row_num, 8).value),
                refeicao=self._safe_int(ws.cell(row_num, 10).value),
                repeticao_refeicao=self._safe_int(ws.cell(row_num, 11).value),
                sobremesa=self._safe_int(ws.cell(row_num, 12).value),
                repeticao_sobremesa=self._safe_int(ws.cell(row_num, 13).value),
                refeicao_2a=self._safe_int(ws.cell(row_num, 14).value),
                repeticao_refeicao_2a=self._safe_int(ws.cell(row_num, 15).value),
                sobremesa_2a=self._safe_int(ws.cell(row_num, 16).value),
                repeticao_sobremesa_2a=self._safe_int(ws.cell(row_num, 17).value)
            ))

            # 1º PERÍODO (7 fields) - cols R,T,U,X,AB,AE,AI (18,20,21,24,28,31,35)
            primeiro_periodo.append(DailyFrequencyRecord(
                day=day_int,
                frequencia=self._safe_int(ws.cell(row_num, 18).value),
                lanche_4h=self._safe_int(ws.cell(row_num, 20).value),
                lanche_6h=self._safe_int(ws.cell(row_num, 21).value),
                refeicao=self._safe_int(ws.cell(row_num, 24).value),
                repeticao_refeicao=self._safe_int(ws.cell(row_num, 28).value),
                sobremesa=self._safe_int(ws.cell(row_num, 31).value),
                repeticao_sobremesa=self._safe_int(ws.cell(row_num, 35).value)
            ))

            # INTERMEDIÁRIO (6 fields) - cols AK,AL,AM,AO,AQ,AS (37,38,39,41,43,45)
            intermediario.append(DailyFrequencyIntermediario(
                day=day_int,
                frequencia=self._safe_int(ws.cell(row_num, 37).value),
                lanche_4h=self._safe_int(ws.cell(row_num, 38).value),
                refeicao=self._safe_int(ws.cell(row_num, 39).value),
                repeticao_refeicao=self._safe_int(ws.cell(row_num, 41).value),
                sobremesa=self._safe_int(ws.cell(row_num, 43).value),
                repeticao_sobremesa=self._safe_int(ws.cell(row_num, 45).value)
            ))

            # 3º PERÍODO (7 fields) - cols AU,AW,AY,BE,BI,BJ,BQ (47,49,51,57,61,62,69)
            terceiro_periodo.append(DailyFrequencyRecord(
                day=day_int,
                frequencia=self._safe_int(ws.cell(row_num, 47).value),
                lanche_4h=self._safe_int(ws.cell(row_num, 49).value),
                lanche_6h=self._safe_int(ws.cell(row_num, 51).value),
                refeicao=self._safe_int(ws.cell(row_num, 57).value),
                repeticao_refeicao=self._safe_int(ws.cell(row_num, 61).value),
                sobremesa=self._safe_int(ws.cell(row_num, 62).value),
                repeticao_sobremesa=self._safe_int(ws.cell(row_num, 69).value)
            ))

            # DOCE checkboxes (4 fields) - cols BR,BU,BV,BW (70,73,74,75)
            def is_checked(val) -> Optional[bool]:
                if val is None or val == "":
                    return None
                return str(val).strip().upper() in ('X', 'TRUE', '1', 'YES')

            doce_checkboxes.append(DailyDoceCheckboxes(
                day=day_int,
                integral=is_checked(ws.cell(row_num, 70).value),
                primeiro_periodo=is_checked(ws.cell(row_num, 73).value),
                intermediario=is_checked(ws.cell(row_num, 74).value),
                terceiro_periodo=is_checked(ws.cell(row_num, 75).value)
            ))

        # Extract TOTAL row (row 59)
        total_row = 59
        total = Section2TotalRow(
            # INTEGRAL (11 fields)
            integral_frequencia=self._safe_int(ws.cell(total_row, 5).value),
            integral_lanche_4h=self._safe_int(ws.cell(total_row, 7).value),
            integral_lanche_6h=self._safe_int(ws.cell(total_row, 8).value),
            integral_refeicao=self._safe_int(ws.cell(total_row, 10).value),
            integral_repeticao_refeicao=self._safe_int(ws.cell(total_row, 11).value),
            integral_sobremesa=self._safe_int(ws.cell(total_row, 12).value),
            integral_repeticao_sobremesa=self._safe_int(ws.cell(total_row, 13).value),
            integral_refeicao_2a=self._safe_int(ws.cell(total_row, 14).value),
            integral_repeticao_refeicao_2a=self._safe_int(ws.cell(total_row, 15).value),
            integral_sobremesa_2a=self._safe_int(ws.cell(total_row, 16).value),
            integral_repeticao_sobremesa_2a=self._safe_int(ws.cell(total_row, 17).value),
            # P1 (7 fields)
            p1_frequencia=self._safe_int(ws.cell(total_row, 18).value),
            p1_lanche_4h=self._safe_int(ws.cell(total_row, 20).value),
            p1_lanche_6h=self._safe_int(ws.cell(total_row, 21).value),
            p1_refeicao=self._safe_int(ws.cell(total_row, 24).value),
            p1_repeticao_refeicao=self._safe_int(ws.cell(total_row, 28).value),
            p1_sobremesa=self._safe_int(ws.cell(total_row, 31).value),
            p1_repeticao_sobremesa=self._safe_int(ws.cell(total_row, 35).value),
            # INTERMEDIÁRIO (6 fields)
            intermediario_frequencia=self._safe_int(ws.cell(total_row, 37).value),
            intermediario_lanche_4h=self._safe_int(ws.cell(total_row, 38).value),
            intermediario_refeicao=self._safe_int(ws.cell(total_row, 39).value),
            intermediario_repeticao_refeicao=self._safe_int(ws.cell(total_row, 41).value),
            intermediario_sobremesa=self._safe_int(ws.cell(total_row, 43).value),
            intermediario_repeticao_sobremesa=self._safe_int(ws.cell(total_row, 45).value),
            # P3 (7 fields)
            p3_frequencia=self._safe_int(ws.cell(total_row, 47).value),
            p3_lanche_4h=self._safe_int(ws.cell(total_row, 49).value),
            p3_lanche_6h=self._safe_int(ws.cell(total_row, 51).value),
            p3_refeicao=self._safe_int(ws.cell(total_row, 57).value),
            p3_repeticao_refeicao=self._safe_int(ws.cell(total_row, 61).value),
            p3_sobremesa=self._safe_int(ws.cell(total_row, 62).value),
            p3_repeticao_sobremesa=self._safe_int(ws.cell(total_row, 69).value)
        )

        return Section2Data(
            integral=integral,
            primeiro_periodo=primeiro_periodo,
            intermediario=intermediario,
            terceiro_periodo=terceiro_periodo,
            doce_checkboxes=doce_checkboxes,
            total=total
        )

    def _extract_section3(self, ws) -> Section3Data:
        """Extract Section 3: Daily special diet data (11 fields × 31 days)"""
        # Section 3 starts at row 77, data in columns C(3), D(4), F(6), H(8), K(11), M(13), O(15), Q(17), S(19), U(21), W+(23+)
        DAY_START_ROW = 77
        TOTAL_ROW = 108  # Row 77 + 31 days = row 108

        # Column mapping (Excel 1-indexed)
        COL_DAY = 3  # C
        COL_GRUPO_A_FREQ = 4  # D
        COL_GRUPO_A_LANCHE_4H = 6  # F
        COL_GRUPO_A_LANCHE_6H = 8  # H
        COL_GRUPO_A_REFEICAO_ENTERAL = 11  # K
        COL_GRUPO_B_FREQ = 13  # M
        COL_GRUPO_B_LANCHE_4H = 15  # O
        COL_GRUPO_B_LANCHE_6H = 17  # Q
        COL_LANCHE_EMERGENCIAL = 19  # S
        COL_KIT_LANCHE = 21  # U
        COL_OBSERVACOES_START = 23  # W+ (observations can span multiple columns)

        days = []
        for day in range(1, 32):
            row = DAY_START_ROW + (day - 1)

            # Check for observations (scan columns 23-30 for text)
            observacoes = None
            for col in range(COL_OBSERVACOES_START, COL_OBSERVACOES_START + 8):
                cell_val = ws.cell(row, col).value
                if cell_val and isinstance(cell_val, str) and len(cell_val) > 1:
                    observacoes = str(cell_val).strip()
                    break

            day_data = Section3DayData(
                day=day,
                grupo_a_frequencia=self._safe_int(ws.cell(row, COL_GRUPO_A_FREQ).value),
                grupo_a_lanche_4h=self._safe_int(ws.cell(row, COL_GRUPO_A_LANCHE_4H).value),
                grupo_a_lanche_6h=self._safe_int(ws.cell(row, COL_GRUPO_A_LANCHE_6H).value),
                grupo_a_refeicao_enteral=self._safe_int(ws.cell(row, COL_GRUPO_A_REFEICAO_ENTERAL).value),
                grupo_b_frequencia=self._safe_int(ws.cell(row, COL_GRUPO_B_FREQ).value),
                grupo_b_lanche_4h=self._safe_int(ws.cell(row, COL_GRUPO_B_LANCHE_4H).value),
                grupo_b_lanche_6h=self._safe_int(ws.cell(row, COL_GRUPO_B_LANCHE_6H).value),
                lanche_emergencial=self._safe_int(ws.cell(row, COL_LANCHE_EMERGENCIAL).value),
                kit_lanche=self._safe_int(ws.cell(row, COL_KIT_LANCHE).value),
                observacoes=observacoes
            )
            days.append(day_data)

        # Extract TOTAL row
        total = Section3TotalRow(
            grupo_a_frequencia=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_A_FREQ).value),
            grupo_a_lanche_4h=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_A_LANCHE_4H).value),
            grupo_a_lanche_6h=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_A_LANCHE_6H).value),
            grupo_a_refeicao_enteral=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_A_REFEICAO_ENTERAL).value),
            grupo_b_frequencia=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_B_FREQ).value),
            grupo_b_lanche_4h=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_B_LANCHE_4H).value),
            grupo_b_lanche_6h=self._safe_int(ws.cell(TOTAL_ROW, COL_GRUPO_B_LANCHE_6H).value),
            lanche_emergencial=self._safe_int(ws.cell(TOTAL_ROW, COL_LANCHE_EMERGENCIAL).value),
            kit_lanche=self._safe_int(ws.cell(TOTAL_ROW, COL_KIT_LANCHE).value)
        )

        return Section3Data(days=days, total=total)

    def _safe_int(self, value) -> Optional[int]:
        """Safely convert value to int"""
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None


# ============================================================================
# USAGE EXAMPLE
# ============================================================================

if __name__ == "__main__":
    parser = CustomExcelParser()
    data = parser.parse_excel("/Users/writetodennis/dev/SME/019382.xlsm")

    print(f"EMEI Code: {data.header.emei_code}")
    print(f"EMEI Name: {data.header.emei_name}")
    print(f"\nEnrollment:")
    for period in data.section1.periods:
        print(f"  {period.period_name}: {period.num_students} students, "
              f"Diet A: {period.special_diet_a}, Diet B: {period.special_diet_b}")
    print(f"  TOTAL: {data.section1.total_students} students")

    print(f"\nDaily Frequency (first 3 days, 1º PERÍODO):")
    for day_rec in data.section2.primeiro_periodo[:3]:
        print(f"  Day {day_rec.day}: Freq={day_rec.frequencia}, Lanche={day_rec.lanche}, Ref={day_rec.refeicao}")
