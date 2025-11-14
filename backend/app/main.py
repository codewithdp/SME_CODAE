"""
FastAPI Application for Reconciliation System
Using FastAPI Background Tasks (NO Celery/Redis needed!)
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import uuid
import os
from datetime import datetime
import logging
import asyncio
import shutil
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Database
from sqlalchemy import create_engine, Column, String, Integer, Float, Boolean, DateTime, JSON, Text, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.dialects.postgresql import UUID

# Azure Blob Storage (optional)
try:
    from azure.storage.blob import BlobServiceClient
    AZURE_STORAGE_AVAILABLE = True
except ImportError:
    AZURE_STORAGE_AVAILABLE = False

# Our modules
from .excel_parser_custom import CustomExcelParser as ExcelParser
from .pdf_processor import PDFProcessor
from .reconciliation_engine_comprehensive import ComprehensiveReconciliationEngine as ReconciliationEngine, ReconciliationResult

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================

class Settings(BaseModel):
    # Database
    DATABASE_URL: str = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/reconciliation")
    
    # Azure Storage (optional - can use local file system)
    AZURE_STORAGE_CONNECTION_STRING: str = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
    AZURE_STORAGE_CONTAINER: str = os.getenv("AZURE_STORAGE_CONTAINER", "reconciliation-files")
    
    # Azure Document Intelligence
    AZURE_DI_ENDPOINT: str = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT", "")
    AZURE_DI_KEY: str = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY", "")
    
    # Processing
    MIN_PDF_CONFIDENCE: float = float(os.getenv("MIN_PDF_CONFIDENCE", "0.75"))
    MAX_FILE_SIZE_MB: int = int(os.getenv("MAX_FILE_SIZE_MB", "50"))
    
    # File storage
    UPLOAD_DIR: str = os.getenv("UPLOAD_DIR", "/tmp/reconciliation_uploads")
    REPORTS_DIR: str = os.getenv("REPORTS_DIR", "/tmp/reconciliation_reports")

settings = Settings()

# Create upload directories
os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
os.makedirs(settings.REPORTS_DIR, exist_ok=True)

# ============================================================================
# DATABASE SETUP
# ============================================================================

Base = declarative_base()
engine = create_engine(settings.DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

class ReconciliationDB(Base):
    __tablename__ = "reconciliations"
    
    id = Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    emei_id = Column(String(50), index=True)
    excel_filename = Column(String(255), nullable=False)
    pdf_filename = Column(String(255), nullable=False)
    
    # File paths (local storage)
    excel_file_path = Column(String(500))
    pdf_file_path = Column(String(500))
    
    # Azure Blob URLs (optional)
    excel_blob_url = Column(String(500))
    pdf_blob_url = Column(String(500))
    
    # Status
    status = Column(String(20), nullable=False, index=True, default="processing")
    created_at = Column(DateTime, default=datetime.utcnow, index=True)
    completed_at = Column(DateTime, nullable=True)
    
    # Progress tracking
    progress_percentage = Column(Integer, default=0)
    current_step = Column(String(100), default="Starting...")
    
    # Results
    id_match = Column(Boolean)
    pdf_confidence_ok = Column(Boolean)
    total_mismatches = Column(Integer)
    overall_match_percentage = Column(Float)
    
    # Row counts
    excel_row_count = Column(Integer)
    pdf_row_count = Column(Integer)
    row_count_match = Column(Boolean)
    
    # Error tracking
    error_message = Column(Text)
    
    # Full result JSON
    result_data = Column(JSON)
    
    # Report paths
    excel_report_path = Column(String(500))
    pdf_report_path = Column(String(500))

# Create tables
Base.metadata.create_all(bind=engine)

# ============================================================================
# FASTAPI APP
# ============================================================================

app = FastAPI(
    title="Excel-PDF Reconciliation API",
    description="API for reconciling Excel files with their PDF counterparts (using FastAPI Background Tasks)",
    version="2.0.0"
)

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# AZURE BLOB STORAGE CLIENT (Optional)
# ============================================================================

blob_service_client = None
if AZURE_STORAGE_AVAILABLE and settings.AZURE_STORAGE_CONNECTION_STRING:
    try:
        blob_service_client = BlobServiceClient.from_connection_string(
            settings.AZURE_STORAGE_CONNECTION_STRING
        )
        logger.info("Azure Blob Storage initialized")
    except Exception as e:
        logger.warning(f"Azure Blob Storage not available: {e}")

# ============================================================================
# REQUEST/RESPONSE MODELS
# ============================================================================

class UploadResponse(BaseModel):
    reconciliation_id: str
    status: str
    message: str

class StatusResponse(BaseModel):
    reconciliation_id: str
    status: str
    progress_percentage: int
    current_step: str
    error_message: Optional[str] = None
    result: Optional[dict] = None

class ReconciliationListItem(BaseModel):
    reconciliation_id: str
    emei_id: Optional[str]
    excel_filename: str
    pdf_filename: str
    status: str
    created_at: datetime
    completed_at: Optional[datetime]
    overall_match_percentage: Optional[float]
    total_mismatches: Optional[int]

class ReconciliationListResponse(BaseModel):
    items: List[ReconciliationListItem]
    total: int
    page: int
    pages: int

# ============================================================================
# BACKGROUND PROCESSING FUNCTION
# ============================================================================

def process_reconciliation_background(
    reconciliation_id: str,
    excel_path: str,
    pdf_path: str
):
    """
    Background task to process reconciliation
    This runs asynchronously without blocking the API response
    """
    db = SessionLocal()
    
    try:
        logger.info(f"Starting background processing for {reconciliation_id}")
        
        # Update status: Starting
        update_progress(db, reconciliation_id, 5, "Validating files...")
        
        # Validate files exist
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        # Update status: Parsing Excel
        update_progress(db, reconciliation_id, 15, "Parsing Excel file...")
        
        # Parse Excel
        excel_parser = ExcelParser()
        excel_data = excel_parser.parse_file(excel_path)
        
        logger.info(f"Excel parsed successfully for {reconciliation_id}")
        
        # Update status: Processing PDF
        update_progress(db, reconciliation_id, 40, "Processing PDF with Azure Document Intelligence...")
        
        # Process PDF
        if not settings.AZURE_DI_ENDPOINT or not settings.AZURE_DI_KEY:
            raise ValueError("Azure Document Intelligence credentials not configured")
        
        pdf_processor = PDFProcessor(
            settings.AZURE_DI_ENDPOINT,
            settings.AZURE_DI_KEY,
            settings.MIN_PDF_CONFIDENCE
        )
        
        # Run async PDF processing in event loop
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        pdf_data = loop.run_until_complete(pdf_processor.process_pdf(pdf_path))
        loop.close()
        
        logger.info(f"PDF processed successfully for {reconciliation_id}")
        
        # Update status: Reconciling
        update_progress(db, reconciliation_id, 70, "Comparing Excel and PDF data...")
        
        # Run reconciliation
        engine = ReconciliationEngine(settings.MIN_PDF_CONFIDENCE)
        result = engine.reconcile(excel_data, pdf_data, reconciliation_id)
        
        logger.info(f"Reconciliation complete for {reconciliation_id}: {result.total_mismatches} mismatches")
        
        # Update status: Generating reports
        update_progress(db, reconciliation_id, 90, "Generating reports...")
        
        # Generate reports (placeholder - implement actual report generation)
        excel_report_path = generate_excel_report(reconciliation_id, result)
        pdf_report_path = None  # Implement if needed

        # Update status: Saving results
        update_progress(db, reconciliation_id, 95, "Saving results...")

        # Close and create fresh session for final save
        db.close()
        db = SessionLocal()

        try:
            # Save final results to database
            reconciliation = db.query(ReconciliationDB).filter(
                ReconciliationDB.id == reconciliation_id
            ).first()

            if reconciliation:
                logger.info(f"Updating reconciliation record for {reconciliation_id}")

                reconciliation.status = "completed"
                reconciliation.completed_at = datetime.utcnow()
                reconciliation.progress_percentage = 100
                reconciliation.current_step = "Complete"

                reconciliation.emei_id = result.emei_id_excel
                reconciliation.id_match = result.id_match
                reconciliation.pdf_confidence_ok = result.pdf_confidence_ok
                reconciliation.total_mismatches = result.total_mismatches
                reconciliation.overall_match_percentage = result.overall_match_percentage
                reconciliation.excel_row_count = result.excel_row_count
                reconciliation.pdf_row_count = result.pdf_row_count
                reconciliation.row_count_match = result.row_count_match

                reconciliation.excel_report_path = excel_report_path
                reconciliation.pdf_report_path = pdf_report_path

                # Save complete result data matching ReconciliationResult model
                reconciliation.result_data = {
                    # IDs and timestamps
                    "reconciliation_id": str(result.reconciliation_id),
                    "timestamp": result.timestamp.isoformat() if hasattr(result.timestamp, 'isoformat') else str(result.timestamp),

                    # ID matching
                    "emei_code_match": result.emei_code_match,
                    "excel_emei": result.excel_emei,
                    "pdf_emei": result.pdf_emei,
                    "emei_id_excel": result.emei_id_excel,
                    "id_match": result.id_match,

                    # PDF quality
                    "pdf_confidence_ok": result.pdf_confidence_ok,
                    "pdf_overall_confidence": float(result.pdf_overall_confidence),

                    # Comparison results
                    "total_mismatches": result.total_mismatches,
                    "total_cells_compared": result.total_cells_compared,
                    "mismatches": [
                        {
                            "section": m.section,
                            "field": m.field,
                            "day_or_period": m.row_identifier,  # Frontend expects this field name
                            "excel_value": m.excel_value,
                            "pdf_value": m.pdf_value,
                        }
                        for m in result.mismatches
                    ],

                    # Summary metrics
                    "excel_total_students": result.excel_total_students,
                    "pdf_total_students": result.pdf_total_students,
                    "excel_row_count": result.excel_row_count,
                    "pdf_row_count": result.pdf_row_count,
                    "row_count_match": result.row_count_match,

                    # File names
                    "excel_filename": result.excel_filename,
                    "pdf_filename": result.pdf_filename,

                    # Overall metrics
                    "overall_match_percentage": float(result.overall_match_percentage),
                    "status": result.status,

                    # Frontend convenience fields
                    "matching_cells": result.total_cells_compared - result.total_mismatches,
                    "mismatching_cells": result.total_mismatches,
                    "match_percentage": float(result.overall_match_percentage),
                    "excel_file": result.excel_filename,
                    "pdf_file": result.pdf_filename,
                }

                logger.info(f"Committing changes for {reconciliation_id}")
                db.commit()
                logger.info(f"Commit successful for {reconciliation_id}")

            logger.info(f"Processing completed successfully for {reconciliation_id}")

        except Exception as save_error:
            logger.error(f"Error saving results for {reconciliation_id}: {save_error}", exc_info=True)
            # Try to at least mark as completed even if result_data fails
            try:
                reconciliation.status = "completed"
                reconciliation.completed_at = datetime.utcnow()
                reconciliation.progress_percentage = 100
                reconciliation.current_step = "Complete"
                db.commit()
            except:
                pass
            raise
        
    except Exception as e:
        logger.error(f"Error processing reconciliation {reconciliation_id}: {e}", exc_info=True)
        
        # Update status to failed
        reconciliation = db.query(ReconciliationDB).filter(
            ReconciliationDB.id == reconciliation_id
        ).first()
        
        if reconciliation:
            reconciliation.status = "failed"
            reconciliation.error_message = str(e)
            reconciliation.current_step = "Failed"
            db.commit()
        
        raise
    
    finally:
        db.close()

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def update_progress(db: Session, reconciliation_id: str, progress: int, step: str):
    """Update progress in database"""
    reconciliation = db.query(ReconciliationDB).filter(
        ReconciliationDB.id == reconciliation_id
    ).first()
    
    if reconciliation:
        reconciliation.progress_percentage = progress
        reconciliation.current_step = step
        db.commit()
        logger.info(f"{reconciliation_id}: {progress}% - {step}")

def generate_excel_report(reconciliation_id: str, result: ReconciliationResult) -> str:
    """
    Generate comprehensive Excel report with mismatch details
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    report_path = os.path.join(settings.REPORTS_DIR, f"{reconciliation_id}_report.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws['A1'] = "Reconciliation Report"
    ws['A1'].font = Font(size=16, bold=True)

    # Summary section
    ws['A3'] = "Reconciliation ID:"
    ws['B3'] = str(result.reconciliation_id)

    ws['A4'] = "EMEI ID:"
    ws['B4'] = result.emei_id_excel

    ws['A5'] = "Excel File:"
    ws['B5'] = result.excel_filename

    ws['A6'] = "PDF File:"
    ws['B6'] = result.pdf_filename

    ws['A7'] = "Status:"
    ws['B7'] = result.status.upper()
    status_fill = PatternFill(start_color="90EE90" if result.status == "match" else "FFB6C1",
                               end_color="90EE90" if result.status == "match" else "FFB6C1",
                               fill_type="solid")
    ws['B7'].fill = status_fill

    ws['A9'] = "Total Cells Compared:"
    ws['B9'] = result.total_cells_compared
    ws['B9'].font = Font(bold=True)

    ws['A10'] = "Total Mismatches:"
    ws['B10'] = result.total_mismatches
    ws['B10'].font = Font(bold=True, color="FF0000" if result.total_mismatches > 0 else "000000")

    ws['A11'] = "Match Percentage:"
    ws['B11'] = f"{result.overall_match_percentage:.2f}%"
    ws['B11'].font = Font(bold=True, size=12)

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    # Create Mismatches sheet if there are any
    if result.mismatches and len(result.mismatches) > 0:
        ws_mismatches = wb.create_sheet("Mismatches")

        # Headers
        headers = ["Section", "Field", "Row/Period", "Excel Value", "PDF Value", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws_mismatches.cell(1, col, header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for row_idx, mismatch in enumerate(result.mismatches, 2):
            ws_mismatches.cell(row_idx, 1, mismatch.section)
            ws_mismatches.cell(row_idx, 2, mismatch.field)
            ws_mismatches.cell(row_idx, 3, mismatch.row_identifier)
            ws_mismatches.cell(row_idx, 4, str(mismatch.excel_value))
            ws_mismatches.cell(row_idx, 5, str(mismatch.pdf_value))
            ws_mismatches.cell(row_idx, 6, mismatch.description)

        # Adjust column widths
        for col in range(1, 7):
            ws_mismatches.column_dimensions[chr(64 + col)].width = 20
        ws_mismatches.column_dimensions['F'].width = 50

    # Save
    wb.save(report_path)
    logger.info(f"Report generated: {report_path} with {len(result.mismatches)} mismatches detailed")

    return report_path

# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/")
def root():
    return {
        "message": "Excel-PDF Reconciliation API",
        "version": "2.0.0 (Background Tasks)",
        "status": "running"
    }

@app.get("/health")
def health_check():
    """Health check endpoint"""
    db_status = "unknown"

    try:
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
        db_status = "healthy"
    except Exception as e:
        db_status = f"unhealthy: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "healthy" else "degraded",
        "timestamp": datetime.utcnow().isoformat(),
        "database": db_status,
        "azure_di_configured": bool(settings.AZURE_DI_ENDPOINT and settings.AZURE_DI_KEY)
    }

@app.post("/api/v1/reconciliation/upload", response_model=UploadResponse)
async def upload_files(
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...),
    pdf_file: UploadFile = File(...)
):
    """
    Upload Excel and PDF files for reconciliation
    Returns immediately and processes in background
    """
    try:
        # Validate file types
        if not excel_file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
            raise HTTPException(status_code=400, detail="Invalid Excel file format. Expected .xlsx, .xlsm, or .xls")
        
        if not pdf_file.filename.endswith('.pdf'):
            raise HTTPException(status_code=400, detail="Invalid PDF file format. Expected .pdf")
        
        # Check file sizes
        max_size = settings.MAX_FILE_SIZE_MB * 1024 * 1024
        
        # Generate reconciliation ID
        reconciliation_id = str(uuid.uuid4())
        
        # Create directory for this reconciliation
        rec_dir = os.path.join(settings.UPLOAD_DIR, reconciliation_id)
        os.makedirs(rec_dir, exist_ok=True)
        
        # Save uploaded files
        excel_path = os.path.join(rec_dir, excel_file.filename)
        pdf_path = os.path.join(rec_dir, pdf_file.filename)
        
        # Save Excel file
        with open(excel_path, "wb") as f:
            content = await excel_file.read()
            if len(content) > max_size:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Excel file too large (max {settings.MAX_FILE_SIZE_MB}MB)"
                )
            f.write(content)
        
        # Save PDF file
        with open(pdf_path, "wb") as f:
            content = await pdf_file.read()
            if len(content) > max_size:
                raise HTTPException(
                    status_code=400, 
                    detail=f"PDF file too large (max {settings.MAX_FILE_SIZE_MB}MB)"
                )
            f.write(content)
        
        logger.info(f"Files saved for reconciliation {reconciliation_id}")
        
        # Upload to Azure Blob Storage (optional)
        excel_blob_url = None
        pdf_blob_url = None
        
        if blob_service_client:
            try:
                container_client = blob_service_client.get_container_client(
                    settings.AZURE_STORAGE_CONTAINER
                )
                
                # Upload Excel
                excel_blob_name = f"excel/{reconciliation_id}/{excel_file.filename}"
                blob_client = container_client.get_blob_client(excel_blob_name)
                with open(excel_path, "rb") as data:
                    blob_client.upload_blob(data, overwrite=True)
                excel_blob_url = blob_client.url
                
                # Upload PDF
                pdf_blob_name = f"pdf/{reconciliation_id}/{pdf_file.filename}"
                blob_client = container_client.get_blob_client(pdf_blob_name)
                with open(pdf_path, "rb") as data:
                    blob_client.upload_blob(data, overwrite=True)
                pdf_blob_url = blob_client.url
                
                logger.info(f"Files uploaded to Azure Blob Storage for {reconciliation_id}")
            except Exception as e:
                logger.warning(f"Failed to upload to Azure Blob Storage: {e}")
        
        # Create database record
        db = SessionLocal()
        try:
            db_reconciliation = ReconciliationDB(
                id=reconciliation_id,
                excel_filename=excel_file.filename,
                pdf_filename=pdf_file.filename,
                excel_file_path=excel_path,
                pdf_file_path=pdf_path,
                excel_blob_url=excel_blob_url,
                pdf_blob_url=pdf_blob_url,
                status="processing",
                progress_percentage=0,
                current_step="Queued for processing..."
            )
            db.add(db_reconciliation)
            db.commit()
            logger.info(f"Database record created for {reconciliation_id}")
        finally:
            db.close()
        
        # Add background task (this is the magic - no Celery needed!)
        background_tasks.add_task(
            process_reconciliation_background,
            reconciliation_id,
            excel_path,
            pdf_path
        )
        
        logger.info(f"Background task queued for {reconciliation_id}")
        
        return UploadResponse(
            reconciliation_id=reconciliation_id,
            status="processing",
            message="Files uploaded successfully. Processing started in background."
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading files: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.get("/api/v1/reconciliation/{reconciliation_id}/status", response_model=StatusResponse)
def get_reconciliation_status(reconciliation_id: str):
    """
    Get status of a reconciliation
    Frontend should poll this endpoint every few seconds
    """
    db = SessionLocal()
    try:
        reconciliation = db.query(ReconciliationDB).filter(
            ReconciliationDB.id == reconciliation_id
        ).first()
        
        if not reconciliation:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        
        # Return result_data as-is (skip Pydantic validation)
        result = reconciliation.result_data if reconciliation.result_data else None
        
        return StatusResponse(
            reconciliation_id=reconciliation_id,
            status=reconciliation.status,
            progress_percentage=reconciliation.progress_percentage or 0,
            current_step=reconciliation.current_step or "Processing...",
            error_message=reconciliation.error_message,
            result=result
        )
        
    finally:
        db.close()

@app.get("/api/v1/reconciliation/{reconciliation_id}/report")
def download_report(reconciliation_id: str, format: str = "excel"):
    """
    Download reconciliation report
    """
    db = SessionLocal()
    try:
        reconciliation = db.query(ReconciliationDB).filter(
            ReconciliationDB.id == reconciliation_id
        ).first()
        
        if not reconciliation:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        
        if reconciliation.status != "completed":
            raise HTTPException(status_code=400, detail="Reconciliation not yet completed")
        
        if format == "excel":
            if not reconciliation.excel_report_path or not os.path.exists(reconciliation.excel_report_path):
                raise HTTPException(status_code=404, detail="Excel report not found")
            
            return FileResponse(
                reconciliation.excel_report_path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"reconciliation_{reconciliation_id}_report.xlsx"
            )
        elif format == "pdf":
            if not reconciliation.pdf_report_path or not os.path.exists(reconciliation.pdf_report_path):
                raise HTTPException(status_code=404, detail="PDF report not yet implemented")
            
            return FileResponse(
                reconciliation.pdf_report_path,
                media_type="application/pdf",
                filename=f"reconciliation_{reconciliation_id}_report.pdf"
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")
    
    finally:
        db.close()

@app.get("/api/v1/reconciliations", response_model=ReconciliationListResponse)
def list_reconciliations(page: int = 1, limit: int = 20):
    """
    List all reconciliations with pagination
    """
    if page < 1:
        page = 1
    if limit < 1 or limit > 100:
        limit = 20
    
    db = SessionLocal()
    try:
        # Get total count
        total = db.query(ReconciliationDB).count()
        
        # Get paginated results
        offset = (page - 1) * limit
        reconciliations = db.query(ReconciliationDB).order_by(
            ReconciliationDB.created_at.desc()
        ).offset(offset).limit(limit).all()
        
        items = [
            ReconciliationListItem(
                reconciliation_id=str(r.id),
                emei_id=r.emei_id,
                excel_filename=r.excel_filename,
                pdf_filename=r.pdf_filename,
                status=r.status,
                created_at=r.created_at,
                completed_at=r.completed_at,
                overall_match_percentage=r.overall_match_percentage,
                total_mismatches=r.total_mismatches
            )
            for r in reconciliations
        ]
        
        pages = (total + limit - 1) // limit
        
        return ReconciliationListResponse(
            items=items,
            total=total,
            page=page,
            pages=pages
        )
        
    finally:
        db.close()

@app.delete("/api/v1/reconciliation/{reconciliation_id}")
def delete_reconciliation(reconciliation_id: str):
    """
    Delete a reconciliation and its associated files
    """
    db = SessionLocal()
    try:
        reconciliation = db.query(ReconciliationDB).filter(
            ReconciliationDB.id == reconciliation_id
        ).first()
        
        if not reconciliation:
            raise HTTPException(status_code=404, detail="Reconciliation not found")
        
        # Delete files
        rec_dir = os.path.join(settings.UPLOAD_DIR, reconciliation_id)
        if os.path.exists(rec_dir):
            shutil.rmtree(rec_dir)
        
        if reconciliation.excel_report_path and os.path.exists(reconciliation.excel_report_path):
            os.remove(reconciliation.excel_report_path)
        
        # Delete from database
        db.delete(reconciliation)
        db.commit()
        
        return {"message": "Reconciliation deleted successfully"}
        
    finally:
        db.close()

# ============================================================================
# RUN SERVER
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    
    # Check configuration
    if not settings.AZURE_DI_ENDPOINT or not settings.AZURE_DI_KEY:
        logger.warning("⚠️  Azure Document Intelligence not configured!")
        logger.warning("Set AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT and AZURE_DOCUMENT_INTELLIGENCE_KEY")
    
    logger.info("🚀 Starting Reconciliation API (Background Tasks mode)")
    logger.info(f"📁 Upload directory: {settings.UPLOAD_DIR}")
    logger.info(f"📊 Reports directory: {settings.REPORTS_DIR}")
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
