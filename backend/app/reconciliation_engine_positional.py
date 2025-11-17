"""
Positional Reconciliation Engine

Uses a fixed one-time column mapping between visible Excel columns and PDF columns.
This is based on the fact that the PDF is a printed/scanned version of the Excel file.
"""
import os
import logging
from typing import Dict, List, Optional, Tuple, Any
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)


# Section 2 (Table 2) Excel → PDF column mapping
# Built by analyzing merged cells and skipping hidden/metadata/empty columns
SECTION2_EXCEL_TO_PDF_MAPPING = {
    4: 1,  # E: Frequência (INTEGRAL)
    6: 2,  # G: Lanche (4h)
    7: 3,  # H: Lanche (6h)
    9: 4,  # J: Refeição
    10: 5,  # K: Repetição da refeição
    11: 6,  # L: Sobremesa
    12: 7,  # M: Repetição da sobremesa
    13: 8,  # N: 2a Refeição
    14: 9,  # O: Repetição da  2a refeição
    15: 10,  # P: 2a Sobremesa
    16: 11,  # Q: 2a Repetição da sobremesa
    17: 12,  # R: Frequência (1º PERÍODO)
    19: 13,  # T: Lanche (4h)
    20: 14,  # U: Lanche (6h)
    23: 15,  # X: Refeição
    27: 16,  # AB: Repetição da refeição
    30: 17,  # AE: Sobremesa
    34: 18,  # AI: Repetição da sobremesa
    36: 19,  # AK: Frequência (INTERMEDIARIO)
    37: 20,  # AL: Lanche (4h)
    38: 21,  # AM: Refeição
    40: 22,  # AO: Repetição da refeição
    42: 23,  # AQ: Sobremesa
    44: 24,  # AS: Repetição da sobremesa
    46: 25,  # AU: Frequência (3º PERÍODO)
    48: 26,  # AW: Lanche(4h)
    50: 27,  # AY: Lanche (6h)
    56: 28,  # BE: Refeição
    60: 29,  # BI: Repetição da refeição
    61: 30,  # BJ: Sobremesa
    68: 31,  # BQ: Repetição da sobremesa
    69: 32,  # BR: INTEGRAL (checkbox)
    72: 33,  # BU: 1º PERÍODO (checkbox)
    73: 34,  # BV: INTERMEDIÁRIO (checkbox)
    74: 35,  # BW: 3º PERÍODO (checkbox)
}

# Section 1 (Table 0) Excel → PDF column mapping
# Table 0: 6 rows × 5 columns
# Col 0 = PERÍODOS, Col 1 = Horas, Cols 2-4 = Student counts (reconcile these 3 only)
SECTION1_EXCEL_TO_PDF_MAPPING = {
    11: 2,  # L: Nº DE ALUNOS MATRICULADOS
    17: 3,  # R: Nº DE ALUNOS MATRICULADOS COM DIETA ESPECIAL - A
    21: 4,  # V: Nº DE ALUNOS MATRICULADOS COM DIETA ESPECIAL - B
}

# Section 3 (Table 4) Excel → PDF column mapping
# Table 4: 35 rows × 10 columns (PDF cols 0-9)
# Col 0 = Day number, Cols 1-7 = Data, Cols 8-9 = LANCHE EMERGENCIAL/KIT LANCHE (not in Excel)
SECTION3_EXCEL_TO_PDF_MAPPING = {
    3: 1,   # D: FREQUENCIA (GRUPO A)
    5: 2,   # F: Lanche (4h) (GRUPO A)
    7: 3,   # H: Lanche (6h) (GRUPO A)
    10: 4,  # K: REFEIÇÃO (SOMENTE DIETA ENTERAL) (GRUPO A)
    12: 5,  # M: FREQUENCIA (GRUPO B)
    14: 6,  # O: Lanche (4h) (GRUPO B)
    16: 7,  # Q: Lanche (6h) (GRUPO B)
}


class PositionalReconciliationEngine:
    """Reconciliation engine using fixed positional column mapping"""

    def __init__(self):
        self.endpoint = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT")
        self.key = os.getenv("AZURE_DOCUMENT_INTELLIGENCE_KEY")
        self.model_id = "SEM_EMEI_v1"

        if not self.endpoint or not self.key:
            raise ValueError("Azure Document Intelligence credentials not found in environment")

        self.client = DocumentIntelligenceClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.key)
        )

    @staticmethod
    def excel_column_letter(col_idx: int) -> str:
        """Convert column index (0-based) to Excel column letter"""
        result = ""
        col_idx += 1  # Convert to 1-based
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + ord('A')) + result
            col_idx //= 26
        return result

    def extract_table(self, pdf_path: str, table_index: int = 2) -> Tuple[Any, Dict]:
        """
        Extract a table from PDF using Azure DI
        Args:
            pdf_path: Path to PDF file
            table_index: Index of table to extract (default 2 for Section2)
        Returns:
            (table_object, metadata)
        """
        with open(pdf_path, "rb") as f:
            poller = self.client.begin_analyze_document(
                self.model_id,
                analyze_request=f,
                content_type="application/pdf"
            )
            result = poller.result()

        if not result.tables or len(result.tables) <= table_index:
            raise ValueError(f"Table {table_index} not found in PDF")

        table = result.tables[table_index]

        metadata = {
            "row_count": table.row_count,
            "column_count": table.column_count,
            "table_index": table_index
        }

        return table, metadata

    def build_pdf_table_structure(self, table: Any, data_start_row: int = 2) -> Dict:
        """
        Build a structured representation of the PDF table
        Args:
            table: PDF table object
            data_start_row: Row index where data starts (default 2 for Section2/3, 1 for Section1)
        Returns: {
            "rows": [
                {"row_idx": idx, "day": day_num, "cells": {col_idx: content}},
                ..
            ]
        }
        """
        # Extract data rows (skip header rows)
        rows = []
        for row_idx in range(data_start_row, table.row_count):
            row_data = {"row_idx": row_idx, "cells": {}}

            # Get all cells for this row
            for cell in table.cells:
                if cell.row_index == row_idx:
                    row_data["cells"][cell.column_index] = cell.content if cell.content else ""

            # Get day number from column 0
            day_val = row_data["cells"].get(0, "").strip()
            if day_val and day_val.lower() != "total":
                try:
                    row_data["day"] = int(day_val)
                except (ValueError, AttributeError):
                    row_data["day"] = day_val
            elif "total" in day_val.lower():
                row_data["day"] = "Total"
            else:
                row_data["day"] = None

            rows.append(row_data)

        return {"rows": rows}

    def load_excel_sheet(self, excel_path: str, sheet_name: str = "EMEI") -> Any:
        """Load Excel sheet"""
        wb = load_workbook(excel_path, data_only=True)
        ws = None

        for sn in wb.sheetnames:
            if sheet_name.upper() in sn.upper():
                ws = wb[sn]
                break

        if not ws:
            raise ValueError(f"Sheet containing '{sheet_name}' not found in Excel file")

        return ws

    @staticmethod
    def _normalize_value(value: str) -> str:
        """
        Normalize value for comparison

        Handles:
        1. Empty cells vs :unselected: (both mean unchecked/empty)
        2. Checkbox representations (X/x vs :selected:)
        3. Number formatting (remove thousands separators)
        """
        if not value or value == "":
            return ""

        # Treat :unselected: as empty (unchecked checkbox = empty cell)
        if value == ":unselected:":
            return ""

        # Normalize checkbox values
        if value.upper() == "X":
            return ":selected:"

        # Handle values that contain both content and checkbox markers
        # e.g., "0 :selected:" or "x :selected:"
        if ":selected:" in value:
            # If it's just "x :selected:" or "X :selected:", normalize to :selected:
            cleaned = value.replace(":selected:", "").strip()
            if cleaned.upper() == "X" or cleaned == "":
                return ":selected:"
            # Otherwise keep the value with :selected:
            return value

        # Remove thousands separators from numbers in PDF (1.665 → 1665)
        if "." in value and value.replace(".", "").replace(",", "").isdigit():
            # Check if it looks like a thousands separator (1.665 not 1.5)
            parts = value.split(".")
            if len(parts) == 2 and len(parts[1]) == 3 and parts[1].isdigit():
                return value.replace(".", "")

        return value

    def reconcile_section(
        self,
        pdf_path: str,
        excel_path: str,
        table_index: int = 2,
        excel_start_row: int = 28,
        column_mapping: Dict[int, int] = None,
        pdf_data_start_row: int = 2,
        excel_row_skip: int = 0
    ) -> Dict:
        """
        Reconcile a section between PDF table and Excel using positional mapping

        Args:
            pdf_path: Path to PDF file
            excel_path: Path to Excel file
            table_index: Index of table in PDF (0 for Section1, 2 for Section2, 4 for Section3)
            excel_start_row: Starting row in Excel (1-indexed) for first data day
            column_mapping: Dict mapping Excel col idx to PDF col idx
            pdf_data_start_row: PDF row index where Day 1 data starts (1 for Section1, 2 for Section2, 3 for Section3)
            excel_row_skip: Extra offset for Total row (1 for Section1 due to empty row 19)

        Returns:
            Dictionary with reconciliation results
        """
        if column_mapping is None:
            column_mapping = SECTION2_EXCEL_TO_PDF_MAPPING

        # Extract PDF table
        pdf_table, pdf_metadata = self.extract_table(pdf_path, table_index)
        pdf_structure = self.build_pdf_table_structure(pdf_table, data_start_row=pdf_data_start_row)

        # Load Excel
        ws = self.load_excel_sheet(excel_path)

        logger.info(f"Using fixed column mapping: {len(column_mapping)} columns mapped")

        # Reconciliation results
        results = {
            "table_index": table_index,
            "pdf_metadata": pdf_metadata,
            "excel_start_row": excel_start_row,
            "column_mapping": {f"Excel_{k}": f"PDF_{v}" for k, v in column_mapping.items()},
            "days_compared": 0,
            "cells_compared": 0,
            "matches": 0,
            "mismatches": 0,
            "match_percentage": 0.0,
            "day_results": [],
            "mismatched_days": []
        }

        # Process each PDF row
        for pdf_row in pdf_structure["rows"]:
            if pdf_row["day"] is None:
                continue

            day_num = pdf_row["day"]

            # Calculate Excel row for this day
            if isinstance(day_num, int):
                excel_row_idx = excel_start_row + (day_num - 1)
            else:
                # For Total or other special rows, use positional matching
                # pdf_data_start_row is where Day 1 starts in PDF (row 2 for Section2, row 3 for Section3)
                excel_row_idx = excel_start_row + (pdf_row["row_idx"] - pdf_data_start_row)
                # Add excel_row_skip ONLY for Total row (handles empty rows in Excel like Section1 row 19)
                if isinstance(day_num, str) and day_num.lower() == "total":
                    excel_row_idx += excel_row_skip

            excel_row = list(ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx, values_only=True))[0]

            # Compare cells using column mapping
            day_matches = 0
            day_mismatches = 0
            day_cells_compared = 0
            mismatched_cells = []

            for excel_col_idx, pdf_col_idx in column_mapping.items():
                # Get Excel value
                excel_value = excel_row[excel_col_idx] if excel_col_idx < len(excel_row) else None
                excel_str = str(excel_value).strip() if excel_value is not None else ""

                # Get PDF value
                pdf_value = pdf_row["cells"].get(pdf_col_idx, "")
                pdf_str = str(pdf_value).strip() if pdf_value else ""

                day_cells_compared += 1

                # Normalize values for comparison
                excel_normalized = self._normalize_value(excel_str)
                pdf_normalized = self._normalize_value(pdf_str)

                # Compare values
                if excel_normalized == pdf_normalized:
                    day_matches += 1
                else:
                    day_mismatches += 1
                    col_letter = self.excel_column_letter(excel_col_idx)
                    excel_cell_ref = f"{col_letter}{excel_row_idx}"

                    mismatched_cells.append({
                        "excel_column": excel_col_idx,
                        "excel_cell_ref": excel_cell_ref,
                        "excel_value": excel_str if excel_str != "" else "(empty)",
                        "pdf_column": pdf_col_idx,
                        "pdf_value": pdf_str if pdf_str != "" else "(empty)",
                        "excel_row": excel_row_idx
                    })

            results["days_compared"] += 1
            results["cells_compared"] += day_cells_compared
            results["matches"] += day_matches
            results["mismatches"] += day_mismatches

            match_pct = (day_matches / day_cells_compared * 100) if day_cells_compared > 0 else 0
            row_label = f"Day {day_num}" if isinstance(day_num, int) else str(day_num)

            day_result = {
                "day": str(day_num),
                "label": row_label,
                "excel_row": excel_row_idx,
                "cells_compared": day_cells_compared,
                "matches": day_matches,
                "mismatches": day_mismatches,
                "match_percentage": round(match_pct, 2),
                "status": "matched" if day_mismatches == 0 else "mismatched"
            }

            if mismatched_cells:
                day_result["mismatched_cells"] = mismatched_cells

            results["day_results"].append(day_result)

            if day_mismatches > 0:
                results["mismatched_days"].append(str(day_num))

        # Calculate overall match percentage
        if results["cells_compared"] > 0:
            results["match_percentage"] = round(
                (results["matches"] / results["cells_compared"]) * 100, 2
            )

        return results
